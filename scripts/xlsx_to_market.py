#!/usr/bin/env python3
"""
Read root data.xlsx (Myanmar ag market lows/highs by date) and emit:
  - packages/core/src/marketData.generated.ts for @agriora/core (workbook rows +
    optional WFP food CSV via scripts/food_market_lib.py)
  - backend/data/rice_data.csv — one daily series from the workbook (rice or fallback row)
    (mid-price from low/high, synthetic March weather + multi-sentence English headlines:
    rice, oil, logistics, policy, weather) for XGBoost training /api/predict/next-day-pct.
  - backend/data/ml_items/<id>.csv — same schema, one file per MARKET_ITEM id (for per-item training).

WFP-style CSV (vegetables/fruits; meat, fish and eggs) may live at
backend/data/food_cleaned.csv or food_prices_cleaned.csv. Without openpyxl, run
`python3 scripts/merge_food_into_market_ts.py` after regenerating the TS header
from this script, or merge food rows manually.

Tip: add more dated survey cells in data.xlsx where columns are sparse — fuller series
improve charts and the derived training CSV; then re-run this script and `npm run ml:retrain`.

The app’s Rice tab uses `RICE_MARKET_ITEMS`: rows whose category/detail match
riceMarket.ts. If none, demo seed data is used and rice_data.csv is skipped.

Usage (from repo root, with venv + openpyxl):
  .venv/bin/pip install openpyxl
  .venv/bin/python scripts/xlsx_to_market.py
  .venv/bin/python scripts/xlsx_to_market.py --no-rice-csv   # TS only; keeps hand-edited backend/data/rice_data.csv
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))
from food_market_lib import food_csv_market_items  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data.xlsx"
OUT = ROOT / "packages" / "core" / "src" / "marketData.generated.ts"
RICE_CSV_OUT = ROOT / "backend" / "data" / "rice_data.csv"
ML_ITEMS_DIR = ROOT / "backend" / "data" / "ml_items"

_RICE_RE = re.compile(r"ဆန်ခွဲ|ပေါင်စပါး|ဧရာဝတီစပါး")


def is_rice_market_row(item_category: str, item_detail: str) -> bool:
    """Match packages/core/src/riceMarket.ts isRiceMarketItemFromSheet."""
    blob = f"{item_category} {item_detail}"
    if "စပါး" in blob:
        return True
    if "ဆန်" in blob:
        return True
    if _RICE_RE.search(blob):
        return True
    return False


def observation_mid(low_v: float | None, high_v: float | None) -> float | None:
    if low_v is not None and high_v is not None:
        return (low_v + high_v) / 2.0
    if low_v is not None:
        return float(low_v)
    if high_v is not None:
        return float(high_v)
    return None


def mid_usable_count(it: dict) -> int:
    return sum(
        1
        for o in it["observations"]
        if observation_mid(o.get("low"), o.get("high")) is not None
    )


def pick_ml_source_item(items: list[dict]) -> tuple[dict, str] | None:
    """
    Training series for backend rice_data.csv:
    1) Rice row with ≥2 survey points (interpolated to calendar days).
    2) Else commodity row with the most survey points (e.g. ဂျုံ when rice cells are empty).
    """
    rice = [it for it in items if is_rice_market_row(it["itemCategory"], it["itemDetails"])]
    rice_ok = sorted(rice, key=mid_usable_count, reverse=True)
    for it in rice_ok:
        if mid_usable_count(it) >= 2:
            return it, "rice_sheet"
    if not items:
        return None
    best = max(items, key=mid_usable_count)
    if mid_usable_count(best) >= 2:
        return best, "commodity_fallback"
    return None


def daily_series_from_observations(observations: list[dict]) -> list[tuple[str, float]]:
    """Piecewise linear mid prices on every calendar day between first and last survey date."""
    pts: list[tuple[date, float]] = []
    for o in sorted(observations, key=lambda x: x["dateIso"]):
        m = observation_mid(o.get("low"), o.get("high"))
        if m is None:
            continue
        raw = str(o["dateIso"]).strip()[:10]
        try:
            di = datetime.strptime(raw, "%Y-%m-%d").date()
        except ValueError:
            continue
        pts.append((di, float(m)))
    if len(pts) < 2:
        return [(p[0].isoformat(), round(p[1], 2)) for p in pts]

    out: list[tuple[str, float]] = []
    for i in range(len(pts) - 1):
        d0, p0 = pts[i]
        d1, p1 = pts[i + 1]
        span = (d1 - d0).days
        if span <= 0:
            if not out or out[-1][0] != d0.isoformat():
                out.append((d0.isoformat(), round(p0, 2)))
            continue
        for k in range(span):
            dd = d0 + timedelta(days=k)
            t = k / span
            pm = p0 + (p1 - p0) * t
            out.append((dd.isoformat(), round(pm, 2)))
    out.append((pts[-1][0].isoformat(), round(pts[-1][1], 2)))
    return out


def pad_march_training_window(daily: list[tuple[str, float]]) -> list[tuple[str, float]]:
    """
    Extend to 1–24 March in the same year as the first price row (flat outside survey span)
    so XGBoost lag-7 has enough rows when the sheet only has a few survey columns.
    """
    if not daily:
        return daily
    pts = sorted(
        (datetime.strptime(ds, "%Y-%m-%d").date(), float(pr)) for ds, pr in daily
    )
    fd, fp = pts[0]
    ld, lp = pts[-1]
    y, mo = fd.year, fd.month
    start = date(y, mo, 1)
    end = date(y, mo, min(24, 31))
    known = {d: p for d, p in pts}
    out: list[tuple[str, float]] = []
    d = start
    while d <= end:
        if d in known:
            pr = known[d]
        elif d < fd:
            pr = fp
        else:
            pr = lp
        out.append((d.isoformat(), round(pr, 2)))
        d += timedelta(days=1)
    return out


def synthetic_seed_daily(periods_iso: list[str]) -> list[tuple[str, float]]:
    """Match packages/core rice demo: one anchor per sheet column, then interpolate daily."""
    anchors: list[dict] = []
    base_mid = 285_000.0
    volatility = 0.014
    for i, p in enumerate(periods_iso):
        if not p or not str(p).strip():
            continue
        ds = str(p).strip()[:10]
        try:
            datetime.strptime(ds, "%Y-%m-%d")
        except ValueError:
            continue
        drift = 1 + 0.0035 * math.sin(i * 0.85) + i * 0.0011
        mid = base_mid * drift
        w = mid * volatility
        lo, hi = max(0.0, mid - w), mid + w
        anchors.append({"dateIso": ds, "low": lo, "high": hi})
    return daily_series_from_observations(anchors)


# Headlines for ML training (English). Join several per day so sentiment + nf_* features vary.
# Pipeline splits on "." / ";" — use "." inside each sentence only at the end.
_HEADLINE_BANK: tuple[str, ...] = (
    # Rice / grains / food
    "Rice wholesale steady as mandi arrivals pick up after the holiday.",
    "Export demand supports local paddy quotes; traders watch policy signals.",
    "Monsoon outlook improves for delta plantings; futures drift slightly lower.",
    "Logistics delays along main corridors keep wholesale spreads wide.",
    "Bumper harvest chatter caps upside despite tight warehouse stocks.",
    "Heatwave warnings lift irrigation costs; farmgate paddy bids firm.",
    "Regional buyers return after floods; milled rice moves in thin trade.",
    "Government buffer release cools retail rice inflation in urban markets.",
    "Wheat futures rise on Black Sea supply worries and strong import demand.",
    "Corn prices climb as ethanol margins improve and inventories tighten.",
    "Soybean meal demand lifts grain complex; rice follows sentiment higher.",
    "Asia food inflation watch: governments mull release of strategic reserves.",
    "Delta farmers report steady transplant progress; rice quotes unmoved.",
    "Milled rice export registrations pick up ahead of the festival season.",
    "Parboiled rice spreads widen on higher energy and packaging costs.",
    # Oil / energy (nf_oil_energy + price direction)
    "Crude oil prices surge after OPEC signals deeper output cuts next quarter.",
    "Brent futures jump as refinery outages tighten diesel supply in Asia.",
    "Fuel costs rise sharply; truckers warn of higher food distribution charges.",
    "Oil prices fall as inventories build and demand outlook softens.",
    "WTI slides on stronger dollar; energy importers get brief relief.",
    "Diesel prices decline at the pump; logistics firms pass through savings.",
    "OPEC+ meeting ends with surprise production increase; crude eases.",
    "Geopolitical risk premium fades; barrel prices drop for a third session.",
    # Transport / logistics
    "Container freight rates increase on Asia-Europe lanes; ag exporters fret.",
    "Port congestion returns; vessel queues delay bulk grain loadings.",
    "Shipping lines announce surcharges; commodity traders brace for costs.",
    "Freight indices soften as new vessel supply enters the market.",
    "Truck strike threat lifted; inland rice movement normalizes.",
    "Rail bottlenecks ease after ministry intervention; deliveries speed up.",
    # Policy / macro
    "Central bank holds rates; currency stability supports import parity pricing.",
    "Government announces fertilizer subsidy; farmers welcome lower input costs.",
    "Parliament debates export curb on staples; rice traders stay cautious.",
    "Sanctions chatter unsettles commodity finance; spreads widen briefly.",
    "Trade ministry signals tariff review on edible oils; markets react mixed.",
    "IMF mission notes inflation risks from weather and energy pass-through.",
    # Weather / climate
    "Cyclone alert issued; coastal states prepare as rice fields face wind risk.",
    "Monsoon rains arrive early; irrigation demand drops for paddy farmers.",
    "Drought warning in major growing belt; yield fears support grain quotes.",
    "Flood alerts along main rivers; warehouse stocks at risk near ports.",
    "Heatwave persists; power cuts disrupt cold storage for perishables.",
    "La Niña watch raises monsoon uncertainty; ag markets trade nervously.",
    "Normal monsoon forecast lifts planting sentiment; rice prices ease slightly.",
    # Mixed macro / prices (explicit up/down language)
    "Wholesale vegetable prices surge on transport disruption and heat damage.",
    "Retail inflation cools as vegetable basket prices decline month on month.",
    "Commodity index rallies; energy and grains lead the advance.",
    "Agricultural futures ease as harvest pressure hits the cash market.",
    "Import parity calculations rise; domestic rice offers firm up.",
    "Export parity weakens; millers cut offers to attract buyers.",
)


def _composite_headline_for_row(index: int) -> str:
    """Three varied sentences per day (deterministic) for richer training text."""
    n = len(_HEADLINE_BANK)
    if n == 0:
        return "Agricultural markets steady amid thin trade."
    a = (index * 31 + 7) % n
    b = (index * 17 + 3) % n
    c = (index * 13 + 11) % n
    parts = [_HEADLINE_BANK[a], _HEADLINE_BANK[b]]
    if c != a and c != b:
        parts.append(_HEADLINE_BANK[c])
    return " ".join(parts)


def write_daily_ml_csv(
    daily: list[tuple[str, float]], dest: Path, label: str
) -> int:
    """
    One row per calendar day: date, avg_price (mid), rainfall_mm, temp_c, news_headline.
    """
    rows_out: list[tuple[str, float, float, float, str]] = []
    for i, (d_iso, mid) in enumerate(daily):
        rain = round(2.0 + (i * 4.17 + 1.3) % 16.0, 1)
        temp_c = round(29.0 + (i % 9) * 0.75 + 0.15 * (i % 3), 1)
        headline = _composite_headline_for_row(i)
        rows_out.append((d_iso, mid, rain, temp_c, headline))

    dest.parent.mkdir(parents=True, exist_ok=True)
    with dest.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["date", "avg_price", "rainfall_mm", "temp_c", "news_headline"])
        w.writerows(rows_out)
    print(f"  Series: {label} ({len(rows_out)} daily rows)")
    return len(rows_out)


def ts_string(s: str) -> str:
    return json.dumps(s, ensure_ascii=False)


def parse_header_date(cell) -> str:
    if cell is None:
        return ""
    if isinstance(cell, datetime):
        return cell.date().isoformat()
    if isinstance(cell, date):
        return cell.isoformat()
    raw = str(cell).strip()
    raw = re.sub(r"\s+", "", raw)
    raw = re.sub(r"^XX-", "31-", raw, flags=re.IGNORECASE)
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    return raw


def is_number(x) -> bool:
    return isinstance(x, (int, float)) and x is not None


def stable_id(parts: tuple[str, ...]) -> str:
    h = hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()
    return h[:16]


def main() -> None:
    ap = argparse.ArgumentParser(description="Export marketData.generated.ts from data.xlsx")
    ap.add_argument(
        "--no-rice-csv",
        action="store_true",
        help="Do not write backend/data/rice_data.csv",
    )
    ap.add_argument(
        "--rice-csv",
        type=Path,
        default=RICE_CSV_OUT,
        help=f"Output path for ML training CSV (default: {RICE_CSV_OUT})",
    )
    ap.add_argument(
        "--no-ml-items",
        action="store_true",
        help="Do not write per-commodity CSVs under backend/data/ml_items/ (for per-item XGB training)",
    )
    args = ap.parse_args()

    if not XLSX.is_file():
        raise SystemExit(f"Missing {XLSX}")

    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active
    low_cols = [
        c for c in range(1, (ws.max_column or 0) + 1)
        if ws.cell(2, c).value == "Lowest Price"
    ]
    if not low_cols:
        raise SystemExit("Could not find 'Lowest Price' header row")

    periods_iso = [parse_header_date(ws.cell(1, c).value) for c in low_cols]

    filled: list[str | None] = [None, None, None, None, None]
    items: list[dict] = []

    for r in range(3, (ws.max_row or 0) + 1):
        a, b, c, d, e = (ws.cell(r, i).value for i in range(1, 6))
        for i, v in enumerate([a, b, c, d, e]):
            if v is not None and str(v).strip():
                filled[i] = str(v).strip()
        if e is None or not str(e).strip():
            continue

        item_detail = str(e).strip()
        group = filled[0] or ""
        main_category = filled[1] or ""
        category = filled[2] or ""
        item_category = filled[3] or ""

        observations: list[dict] = []
        for idx, lc in enumerate(low_cols):
            lo = ws.cell(r, lc).value
            hi = ws.cell(r, lc + 1).value
            low_v = float(lo) if is_number(lo) else None
            high_v = float(hi) if is_number(hi) else None
            if low_v is None and high_v is None:
                continue
            observations.append(
                {
                    "dateIso": periods_iso[idx],
                    "low": low_v,
                    "high": high_v,
                }
            )

        if not observations:
            continue

        parts = (group, main_category, category, item_category, item_detail, str(r))
        item_id = stable_id(parts)

        items.append(
            {
                "id": item_id,
                "excelRow": r,
                "group": group,
                "mainCategory": main_category,
                "category": category,
                "itemCategory": item_category,
                "itemDetails": item_detail,
                "observations": observations,
            }
        )

    food_items = food_csv_market_items()
    items.extend(food_items)
    if food_items:
        from food_market_lib import FOOD_CSV_CANDIDATES

        used = next((p for p in FOOD_CSV_CANDIDATES if p.is_file()), None)
        if used:
            print(
                f"  Food CSV: {used.name} → {len(food_items)} series (veg/fruit + meat/fish/eggs)"
            )

    generated_at = (
        datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    )

    lines = [
        "/** Auto-generated by scripts/xlsx_to_market.py — do not edit. "
        "Workbook + optional backend/data/food_*cleaned*.csv. */",
        "",
        'import type { MarketItem } from "./marketTypes";',
        "",
        f"export const MARKET_SOURCE_FILE = {ts_string('data.xlsx')};",
        f"export const MARKET_GENERATED_AT_ISO = {ts_string(generated_at)};",
        "",
        "export const MARKET_PERIODS_ISO: readonly string[] = [",
        *[f"  {ts_string(p)}," for p in periods_iso],
        "];",
        "",
        "export const MARKET_ITEMS: readonly MarketItem[] = [",
    ]

    for it in items:
        obs_chunks = []
        for o in it["observations"]:
            low = "null" if o["low"] is None else repr(o["low"])
            high = "null" if o["high"] is None else repr(o["high"])
            obs_chunks.append(
                f"{{ dateIso: {ts_string(o['dateIso'])}, low: {low}, high: {high} }}"
            )
        obs_js = "[" + ", ".join(obs_chunks) + "]"
        lines.append("  {")
        lines.append(f"    id: {ts_string(it['id'])},")
        lines.append(f"    excelRow: {it['excelRow']},")
        lines.append(f"    group: {ts_string(it['group'])},")
        lines.append(f"    mainCategory: {ts_string(it['mainCategory'])},")
        lines.append(f"    category: {ts_string(it['category'])},")
        lines.append(f"    itemCategory: {ts_string(it['itemCategory'])},")
        lines.append(f"    itemDetails: {ts_string(it['itemDetails'])},")
        lines.append(f"    observations: {obs_js},")
        lines.append("  },")

    lines.append("];")
    lines.append("")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {len(items)} items, {len(periods_iso)} periods → {OUT}")

    if args.no_rice_csv:
        return

    picked = pick_ml_source_item(items)
    if picked is not None:
        source, tag = picked
        daily = daily_series_from_observations(source["observations"])
        daily = pad_march_training_window(daily)
        detail = source["itemDetails"][:56]
        print(f"ML training CSV from workbook row: “{detail}…”")
        n = write_daily_ml_csv(daily, args.rice_csv, tag)
    else:
        print("No priced rows in workbook — using synthetic rice series (sheet date columns).")
        daily = synthetic_seed_daily(periods_iso)
        daily = pad_march_training_window(daily)
        n = write_daily_ml_csv(daily, args.rice_csv, "synthetic_seed")

    print(f"→ {args.rice_csv} ({n} rows)")
    print(
        "Retrain: RICE_SENTIMENT_MOCK=1 python backend/train_from_csv.py --csv backend/data/rice_data.csv"
    )

    if not args.no_ml_items:
        ML_ITEMS_DIR.mkdir(parents=True, exist_ok=True)
        n_items = 0
        for it in items:
            if mid_usable_count(it) < 2:
                continue
            daily = daily_series_from_observations(it["observations"])
            daily = pad_march_training_window(daily)
            if len(daily) < 10:
                continue
            item_id = str(it["id"])
            out_csv = ML_ITEMS_DIR / f"{item_id}.csv"
            write_daily_ml_csv(daily, out_csv, item_id)
            n_items += 1
        print(f"→ {ML_ITEMS_DIR} ({n_items} per-item ML CSVs; train: npm run ml:train:items)")


if __name__ == "__main__":
    main()
